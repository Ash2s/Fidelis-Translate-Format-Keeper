import csv
import json
import time
import uuid
import os
import re
from openpyxl import load_workbook
from typing import Optional
from app.config import settings
from app.services.dedup_guard import _is_sentence_like

# How long (in seconds) a glossary is kept since upload.
# 24h was too short for batch workflows — glossaries uploaded once would be
# purged mid-project, silently disabling term enforcement (badcase type 8).
GLOSSARY_TTL = 7 * 24 * 3600  # 7 days

# Header labels to skip when parsing glossary files (CSV/XLSX). The first row
# of a glossary export is often a header like "zh-CN / en-US" or "中文术语 /
# 英文翻译"; loading it as a term pollutes both the prompt injection and the
# replacement pass (observed: '\ufeffzh-CN → en-US' term in real data).
_HEADER_CN = {'中文术语', '中文', '术语', '原文', '中文术语(必填)', '中文名称'}
_HEADER_EN = {'en-us', 'english', '英文', '译文', '翻译', 'translation', 'en', 'en名称'}


class GlossaryService:
    def __init__(self):
        self._glossaries: dict[str, dict[str, str]] = {}
        self._metadata: dict[str, dict] = {}
        # -- Persist glossary data to disk so it survives restarts --
        self._data_dir = settings.GLOSSARY_DIR
        os.makedirs(self._data_dir, exist_ok=True)
        self._load_all_from_disk()

    def _glossary_path(self, glossary_id: str) -> str:
        return os.path.join(self._data_dir, f"{glossary_id}.json")

    def _metadata_path(self, glossary_id: str) -> str:
        return os.path.join(self._data_dir, f"{glossary_id}.meta.json")

    def _save_to_disk(self, glossary_id: str) -> None:
        """Write terms + metadata to disk for this glossary."""
        if glossary_id in self._glossaries:
            with open(self._glossary_path(glossary_id), "w", encoding="utf-8") as f:
                json.dump(self._glossaries[glossary_id], f, ensure_ascii=False)
        if glossary_id in self._metadata:
            meta = dict(self._metadata[glossary_id])
            meta["_uploaded_at"] = time.time()
            with open(self._metadata_path(glossary_id), "w", encoding="utf-8") as f:
                json.dump(meta, f, ensure_ascii=False)

    def _purge_expired(self) -> None:
        """Remove glossaries older than GLOSSARY_TTL."""
        now = time.time()
        expired = []
        for gid, meta in list(self._metadata.items()):
            uploaded = meta.get("_uploaded_at", 0)
            if now - uploaded > GLOSSARY_TTL:
                expired.append(gid)
        for gid in expired:
            self._glossaries.pop(gid, None)
            self._metadata.pop(gid, None)
            for p in (self._glossary_path(gid), self._metadata_path(gid)):
                if os.path.exists(p):
                    try:
                        os.remove(p)
                    except OSError:
                        # Deletion failure (permissions / AV / sandbox) must not
                        # crash service startup. The stale file is simply left
                        # on disk and ignored.
                        pass

    def _load_all_from_disk(self) -> None:
        """Reload every persisted glossary from disk (called at startup)."""
        if not os.path.isdir(self._data_dir):
            return
        for fname in os.listdir(self._data_dir):
            if fname.endswith(".json") and not fname.endswith(".meta.json"):
                gid = fname[:-5]  # strip .json
                try:
                    with open(self._glossary_path(gid), "r", encoding="utf-8") as f:
                        self._glossaries[gid] = json.load(f)
                    meta_path = self._metadata_path(gid)
                    if os.path.exists(meta_path):
                        with open(meta_path, "r", encoding="utf-8") as f:
                            self._metadata[gid] = json.load(f)
                except Exception:
                    continue
        self._purge_expired()

    def load_glossary(self, file_path: str, filename: str) -> str:
        glossary_id = str(uuid.uuid4())
        terms = {}
        ext = os.path.splitext(filename)[-1].lower()
        if ext == ".csv":
            terms = self._load_csv(file_path)
        elif ext in [".xlsx"]:  # openpyxl does not support .xls
            terms = self._load_xlsx(file_path)
        else:
            raise ValueError(f"Unsupported glossary format: {ext}")
        self._glossaries[glossary_id] = terms
        self._metadata[glossary_id] = {
            "filename": filename,
            "term_count": len(terms)
        }
        self._save_to_disk(glossary_id)
        return glossary_id

    def _is_header_row(self, cn: str, en: str) -> bool:
        """True if a parsed row looks like a header rather than a real term.

        Strips UTF-8 BOM and normalizes case/whitespace before comparing.
        """
        cn_norm = cn.strip().lstrip('\ufeff').strip().lower()
        en_norm = en.strip().lower()
        if cn_norm in _HEADER_CN:
            return True
        if en_norm in _HEADER_EN:
            return True
        # "zh-CN" / "en-US" two-column header
        if cn_norm in ('zh-cn', 'zh', 'cn') and en_norm in ('en-us', 'en'):
            return True
        return False

    def _load_csv(self, path: str) -> dict[str, str]:
        terms = {}
        skipped = 0
        with open(path, "r", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            for row in reader:
                if len(row) >= 2:
                    cn, en = row[0].strip(), row[1].strip()
                    if not self._is_header_row(cn, en) and cn and en:
                        if _is_sentence_like(en):
                            skipped += 1
                            continue
                        terms[cn] = en
        if skipped:
            print(f"[glossary] 已忽略 {skipped} 条句子型非规范词条（CSV）")
        return terms

    def _load_xlsx(self, path: str) -> dict[str, str]:
        terms = {}
        skipped = 0
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            if len(row) >= 2:
                if row[0] is not None and row[1] is not None:
                    cn, en = str(row[0]).strip(), str(row[1]).strip()
                    if not self._is_header_row(cn, en) and cn and en:
                        if _is_sentence_like(en):
                            skipped += 1
                            continue
                        terms[cn] = en
        if skipped:
            print(f"[glossary] 已忽略 {skipped} 条句子型非规范词条（XLSX）")
        return terms

    def get_glossary(self, glossary_id: str) -> dict[str, str]:
        if glossary_id not in self._glossaries:
            raise ValueError(f"Glossary not found: {glossary_id}")
        # Touch on access so actively-used glossaries don't expire
        meta = self._metadata.get(glossary_id)
        if meta is not None:
            meta["_uploaded_at"] = time.time()
            with open(self._metadata_path(glossary_id), "w", encoding="utf-8") as f:
                json.dump(meta, f, ensure_ascii=False)
        return self._glossaries[glossary_id]

    def get_term_count(self, glossary_id: str) -> int:
        if glossary_id not in self._glossaries:
            raise ValueError(f"Glossary not found: {glossary_id}")
        return len(self._glossaries[glossary_id])

    def get_metadata(self, glossary_id: str) -> Optional[dict]:
        if glossary_id not in self._metadata:
            raise ValueError(f"Glossary not found: {glossary_id}")
        return self._metadata[glossary_id]

    def normalize_quotes(self, text: str) -> str:
        """Normalize straight double quotes to curly double quotes. Idempotent."""
        # Step 1: normalize any existing curly quotes back to straight (ensures idempotency)
        result = text.replace(chr(0x201C), chr(34)).replace(chr(0x201D), chr(34))
        # Step 2: pair straight quotes — odd positions become left, even become right
        parts = result.split(chr(34))
        result = ""
        for i, part in enumerate(parts):
            if i == len(parts) - 1:
                result += part
            elif i % 2 == 0:
                result += part + chr(0x201C)
            else:
                result += part + chr(0x201D)
        return result

    def replace_with_glossary(self, text: str, glossary: dict[str, str]) -> str:
        """Longest-match-first replacement using glossary dict."""
        if not text or not glossary:
            return text
        # Normalize quotes in text first
        text = self.normalize_quotes(text)
        # Normalize glossary keys for matching
        normalized_glossary = {self.normalize_quotes(k): v for k, v in glossary.items()}
        sorted_terms = sorted(normalized_glossary.keys(), key=len, reverse=True)
        result = text
        for term in sorted_terms:
            pattern = re.escape(term)
            result = re.sub(pattern, normalized_glossary[term], result)
        return result
